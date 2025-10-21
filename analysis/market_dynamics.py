"""Market dynamics analysis toolkit.

This module assembles utilities to:
- load macro, asset, and sentiment time series from the repository datasets
- study macroeconomic factor influence on asset returns via lightweight OLS regressions
- estimate cross-asset lead/lag structure with rolling correlation heuristics
- detect market regimes with an unsupervised clustering pass on returns/volatility
- explore behavioral finance proxies such as volatility and sentiment shocks

Run the module as a script to execute the full workflow using the default
series configuration shipped with the repository.
"""

from __future__ import annotations

import argparse
import logging
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, Optional, Sequence, Tuple
from math import erf

import numpy as np
import pandas as pd


def _normal_cdf(x: np.ndarray) -> np.ndarray:
    """Standard normal cumulative distribution function."""
    x_arr = np.asarray(x, dtype=float)
    scaled = x_arr / np.sqrt(2.0)
    if hasattr(np, "erf"):
        erf_vals = np.erf(scaled)
    else:  # pragma: no cover - fallback for numpy builds without erf
        erf_vals = np.vectorize(erf, otypes=[float])(scaled)
    return 0.5 * (1.0 + erf_vals)


def _ols_fit(y: np.ndarray, X: np.ndarray) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, float]:
    """Return beta, standard error, t-stats, p-values, and R^2 for an OLS regression."""
    XtX = X.T @ X
    try:
        XtX_inv = np.linalg.inv(XtX)
    except np.linalg.LinAlgError:
        XtX_inv = np.linalg.pinv(XtX)
    beta = XtX_inv @ X.T @ y
    resid = y - X @ beta
    dof = max(len(y) - X.shape[1], 1)
    sigma2 = float(resid.T @ resid) / dof
    se = np.sqrt(np.diag(XtX_inv) * sigma2)
    with np.errstate(divide="ignore", invalid="ignore"):
        t_vals = np.divide(beta, se, out=np.full_like(beta, np.nan), where=se != 0)
    p_vals = 2.0 * (1.0 - _normal_cdf(np.abs(t_vals)))
    denom = float(((y - y.mean()) ** 2).sum())
    r_sq = np.nan if denom == 0 else 1.0 - float((resid ** 2).sum()) / denom
    return beta, se, t_vals, p_vals, r_sq


def _rolling_ols(y: pd.Series, X: pd.DataFrame, window: int) -> pd.DataFrame:
    """Compute rolling OLS parameters for a fixed window."""
    if window is None or window <= 0 or window > len(y):
        return pd.DataFrame()
    beta_rows = []
    idx = []
    design_cols = ["const"] + list(X.columns)
    for start in range(0, len(y) - window + 1):
        end = start + window
        y_win = y.iloc[start:end].to_numpy()
        X_win = X.iloc[start:end].to_numpy()
        X_design = np.column_stack([np.ones(len(X_win)), X_win])
        beta, *_ = _ols_fit(y_win, X_design)
        beta_rows.append(beta)
        idx.append(y.index[end - 1])
    return pd.DataFrame(beta_rows, index=idx, columns=design_cols)


def _kmeans(
    data: np.ndarray,
    k: int,
    *,
    n_init: int = 10,
    max_iter: int = 100,
    random_state: int = 42,
) -> np.ndarray:
    """Simple k-means implementation (Euclidean) with multiple initialisations."""
    data = np.asarray(data, dtype=float)
    if data.ndim != 2 or len(data) == 0:
        return np.array([], dtype=int)
    n_samples = data.shape[0]
    if n_samples < k:
        return np.arange(n_samples, dtype=int)
    rng = np.random.default_rng(random_state)
    best_labels: Optional[np.ndarray] = None
    best_inertia = np.inf
    for _ in range(n_init):
        centroids = data[rng.choice(n_samples, size=k, replace=False)]
        labels = np.zeros(n_samples, dtype=int)
        for _ in range(max_iter):
            distances = np.linalg.norm(data[:, None, :] - centroids[None, :, :], axis=2)
            labels = distances.argmin(axis=1)
            new_centroids = np.array(
                [
                    data[labels == j].mean(axis=0) if np.any(labels == j) else centroids[j]
                    for j in range(k)
                ]
            )
            if np.allclose(new_centroids, centroids, atol=1e-6, rtol=0):
                centroids = new_centroids
                break
            centroids = new_centroids
        inertia = float(np.sum((data - centroids[labels]) ** 2))
        if inertia < best_inertia:
            best_inertia = inertia
            best_labels = labels.copy()
    assert best_labels is not None
    return best_labels


def _read_excel_with_openpyxl(path: Path) -> pd.DataFrame:
    """Fallback Excel reader for Bloomberg exports with array formulas."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise ImportError("openpyxl is required to parse Bloomberg Excel exports") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if not any(cell is not None for cell in row):
            continue
        rows.append(list(row))
    wb.close()
    if not rows:
        return pd.DataFrame()

    start_idx = None
    for idx, row in enumerate(rows):
        first = row[0]
        if isinstance(first, (datetime, date)) and first is not None:
            start_idx = idx
            break
    if start_idx is None:
        raise ValueError(f"Could not find a date column in {path}")

    data_rows = rows[start_idx:]
    max_len = max(len(row) for row in data_rows)
    normalized = []
    for row in data_rows:
        padded = row + [None] * (max_len - len(row))
        normalized.append(padded[:max_len])
    columns = [f"col_{idx}" for idx in range(max_len)]
    frame = pd.DataFrame(normalized, columns=columns)
    return frame.replace({"#N/A N/A": np.nan, "#N/A": np.nan})


@dataclass
class SeriesConfig:
    name: str
    path: str
    value_col: Optional[int] = None
    rename: Optional[str] = None
    resample: Optional[str] = None
    transform: Optional[Callable[[pd.Series], pd.Series]] = None

    def resolved_path(self, root: Path) -> Path:
        return root / self.path


def load_excel_series(config: SeriesConfig, data_root: Path) -> pd.Series:
    full_path = config.resolved_path(data_root)
    if not full_path.exists():
        raise FileNotFoundError(f"Missing file for {config.name}: {full_path}")

    try:
        raw = pd.read_excel(full_path)
    except Exception:
        raw = _read_excel_with_openpyxl(full_path)
    if raw.empty:
        raise ValueError(f"File {full_path} has no rows")

    columns = list(raw.columns)
    raw = raw.copy().replace({"#N/A N/A": np.nan, "#N/A": np.nan})
    date_col = columns[0]
    raw[date_col] = pd.to_datetime(raw[date_col], errors="coerce")
    raw = raw.dropna(subset=[date_col])

    data_cols = [col for col in raw.columns[1:] if not raw[col].isna().all()]
    if not data_cols:
        raise ValueError(f"No data columns detected in {full_path}")

    if config.value_col is None:
        target_col = data_cols[-1]
    else:
        idx = config.value_col
        if idx < 0:
            idx = len(data_cols) + idx
        if idx < 0 or idx >= len(data_cols):
            raise IndexError(f"value_col {config.value_col} out of bounds for {config.name}")
        target_col = data_cols[idx]

    series = raw[[date_col, target_col]].dropna()
    series[target_col] = pd.to_numeric(series[target_col], errors="coerce")
    series = series.dropna(subset=[target_col])
    series = series.set_index(date_col)[target_col]
    series = series.sort_index()

    if config.resample:
        series = series.resample(config.resample).last()

    if config.transform:
        series = config.transform(series)

    series.name = config.rename or config.name
    return series


def load_series_frame(configs: Sequence[SeriesConfig], data_root: Path) -> pd.DataFrame:
    series_list = [load_excel_series(cfg, data_root) for cfg in configs]
    frame = pd.concat(series_list, axis=1)
    frame = frame[~frame.index.duplicated(keep="last")]
    frame = frame.sort_index()
    return frame


def compute_log_returns(frame: pd.DataFrame, freq: str = "B") -> pd.DataFrame:
    levels = frame.ffill()
    levels = levels.resample(freq).last()
    returns = np.log(levels).diff()
    returns = returns.dropna(how="all")
    return returns


def compute_feature_changes(frame: pd.DataFrame, freq: str = "B") -> pd.DataFrame:
    levels = frame.ffill()
    levels = levels.resample(freq).last()
    changes = {}
    for col in levels.columns:
        series = levels[col]
        if (series > 0).all():
            change = np.log(series).diff()
        else:
            change = series.diff()
        changes[f"{col}_chg"] = change
    result = pd.DataFrame(changes)
    result = result.dropna(how="all")
    return result


def run_macro_regressions(
    returns: pd.DataFrame,
    macro_features: pd.DataFrame,
    window: Optional[int] = 126,
) -> Tuple[Dict[str, Dict[str, pd.DataFrame]], pd.DataFrame, pd.DataFrame]:
    macro_models: Dict[str, Dict[str, pd.DataFrame]] = {}
    beta_records: list[Dict[str, float]] = []
    corr_records: list[Dict[str, float]] = []
    for asset in returns.columns:
        data = pd.concat([returns[asset], macro_features], axis=1).dropna()
        if data.empty:
            continue
        y = data[asset].to_numpy()
        X_no_const = data.drop(columns=[asset])
        X_design = np.column_stack([np.ones(len(X_no_const)), X_no_const.to_numpy()])
        beta, se, t_vals, p_vals, r_sq = _ols_fit(y, X_design)
        index = ["const"] + list(X_no_const.columns)
        macro_models[asset] = {
            "coef": pd.DataFrame({"coef": beta}, index=index),
            "stderr": pd.DataFrame({"stderr": se}, index=index),
            "tvalue": pd.DataFrame({"tvalue": t_vals}, index=index),
            "pvalue": pd.DataFrame({"pvalue": p_vals}, index=index),
            "rsquared": pd.DataFrame({"rsquared": [r_sq]}),
        }
        for factor, coef_val, err_val, t_val, p_val in zip(
            index[1:], beta[1:], se[1:], t_vals[1:], p_vals[1:]
        ):
            beta_records.append(
                {
                    "asset_class": asset,
                    "macro_factor": factor,
                    "beta": coef_val,
                    "stderr": err_val,
                    "tvalue": t_val,
                    "pvalue": p_val,
                    "r_squared": r_sq,
                }
            )
        corr_series = data.corr().loc[asset].drop(asset, errors="ignore")
        for factor, corr_val in corr_series.items():
            corr_records.append(
                {
                    "asset_class": asset,
                    "macro_factor": factor,
                    "correlation": corr_val,
                }
            )
        if window and window < len(data):
            rolling = _rolling_ols(data[asset], X_no_const, window)
            if not rolling.empty:
                macro_models[asset]["rolling_coef"] = rolling
    macro_beta_table = pd.DataFrame(beta_records)
    macro_corr_table = pd.DataFrame(corr_records)
    return macro_models, macro_beta_table, macro_corr_table


def estimate_cross_asset_influence(
    returns: pd.DataFrame,
    maxlags: int = 5,
) -> Dict[str, pd.DataFrame]:
    records = []
    monthly = returns.resample("M").sum(min_count=1)
    monthly = monthly.dropna(how="all")
    if monthly.empty:
        return {"lead_lag_table": pd.DataFrame(), "top_leads": pd.DataFrame()}

    for source in monthly.columns:
        for target in monthly.columns:
            if source == target:
                continue
            base = pd.concat([monthly[source], monthly[target]], axis=1, join="inner").dropna()
            if base.empty:
                continue
            for lag in range(1, maxlags + 1):
                shifted = base[source].shift(-lag)
                corr = shifted.corr(base[target])
                if pd.notna(corr):
                    records.append(
                        {
                            "leader": source,
                            "follower": target,
                            "lag_months": lag,
                            "lead_correlation": corr,
                        }
                    )

    table = pd.DataFrame(records)
    top = (
        table.sort_values("lead_correlation", ascending=False).head(10)
        if not table.empty
        else pd.DataFrame()
    )
    return {"lead_lag_table": table, "top_leads": top}


def detect_market_regimes(
    returns: pd.Series,
    vol_window: int = 21,
    components: int = 3,
) -> pd.DataFrame:
    df = pd.DataFrame({"ret": returns})
    df["vol"] = returns.rolling(vol_window).std()
    df = df.dropna()

    if df.empty:
        return pd.DataFrame(columns=["ret", "vol", "regime", "regime_label"], index=returns.index)

    features = df.to_numpy(dtype=float)
    mean = np.nanmean(features, axis=0)
    std = np.nanstd(features, axis=0)
    std[std == 0] = 1.0
    features = (features - mean) / std

    labels = _kmeans(features, min(components, len(df)), random_state=42)
    df["regime"] = labels

    mean_returns = df.groupby("regime")["ret"].mean().sort_values()
    ordered_labels = ["bear", "neutral", "bull"]
    mapping = {
        regime: ordered_labels[min(idx, len(ordered_labels) - 1)]
        for idx, regime in enumerate(mean_returns.index)
    }
    df["regime_label"] = df["regime"].map(mapping)
    df = df.reindex(returns.index)
    return df


def behavioral_sentiment_analysis(
    returns: pd.Series,
    sentiment: pd.DataFrame,
    high_z: float = 1.5,
    low_z: float = -1.0,
) -> Dict[str, pd.DataFrame]:
    data = pd.concat([returns, sentiment], axis=1).dropna()
    if data.empty:
        return {"summary": pd.DataFrame(), "z_scores": pd.DataFrame()}

    z_scores = data.apply(lambda col: (col - col.mean()) / col.std(ddof=0))

    metrics = {}
    for col in sentiment.columns:
        series_z = z_scores[col]
        future = returns.shift(-5)  # 1-week lookahead on business days
        joined = pd.concat([series_z, future], axis=1).dropna()
        joined.columns = ["z", "fwd_ret"]

        high = joined[joined["z"] >= high_z]
        low = joined[joined["z"] <= low_z]

        metrics[col] = pd.DataFrame(
            {
                "avg_fwd_ret": [joined["fwd_ret"].mean()],
                "avg_high_sentiment_ret": [high["fwd_ret"].mean() if not high.empty else np.nan],
                "avg_low_sentiment_ret": [low["fwd_ret"].mean() if not low.empty else np.nan],
                "corr_with_ret": [data[col].corr(returns)],
            }
        )

    z_scores.columns = [f"{c}_z" for c in z_scores.columns]
    summary = pd.concat(metrics, axis=0) if metrics else pd.DataFrame()
    return {"summary": summary, "z_scores": z_scores}


def default_config(data_root: Path) -> Dict[str, Sequence[SeriesConfig]]:
    return {
        "assets": [
            SeriesConfig("us_equity", "FULL Nexus Data/Vanguard 500 Index Fund - VFIAX US Equity US.xlsx"),
            SeriesConfig("us_value", "FULL Nexus Data/Vanguard Value Index Fund - VVIAX US Equity US.xlsx"),
            SeriesConfig("us_growth", "FULL Nexus Data/Vanguard Growth Index Fund - VIGAX US Equity US.xlsx"),
            SeriesConfig("balanced", "FULL Nexus Data/Vanguard Balanced Index Fund - VBIAX US Equity US.xlsx"),
            SeriesConfig("global_infra", "FULL Nexus Data/Dow Jones Brookfield Global In - DJBGICUT INDEX Index.xlsx"),
            SeriesConfig("reit", "FULL Nexus Data/FTSE NAREIT All Reits Total Re - FNARTR Index Index.xlsx"),
            SeriesConfig("investment_grade", "FULL Nexus Data/iBoxx USD Liquid Investment Gr - IBOXIG Index Index.xlsx"),
            SeriesConfig("high_yield", "FULL Nexus Data/iBoxx USD Liquid High Yield In - IBOXHY Index Index.xlsx"),
            SeriesConfig("crypto", "FULL Nexus Data/Bloomberg Galaxy Crypto Index - BGCI Index Index.xlsx"),
        ],
        "macro": [
            SeriesConfig("gdp_qoq", "FULL Nexus Data/QoQ % Change Annualized - GDP CQOQ Index Index.xlsx", resample="QS"),
            SeriesConfig("cpi_yoy", "FULL Nexus Data/YoY % NSA - CPI YOY Index Index.xlsx"),
            SeriesConfig("core_cpi_yoy", "FULL Nexus Data/YoY % NSA - CPI XYOY Index Index.xlsx"),
            SeriesConfig("pce_mom", "FULL Nexus Data/Monthly % Change - PCE CRCH Index Index.xlsx"),
            SeriesConfig("ism_pmi", "FULL Nexus Data/ISM PMI - NAPMPMI Index Index.xlsx"),
            SeriesConfig("nfp_change", "FULL Nexus Data/Net Change SA - NFP TCH Index Index.xlsx"),
            SeriesConfig("unemployment", "FULL Nexus Data/Total SA - USURTOT Index Index.xlsx"),
            SeriesConfig("michigan_sent", "FULL Nexus Data/Univ. of Michigan Sentiment - CONSSENT INDEX Index.xlsx"),
            SeriesConfig("conference_board", "FULL Nexus Data/Confidence - CONCCONF INDEX Index.xlsx"),
            SeriesConfig("m2", "FULL Nexus Data/M2 (NSA) - M2NS Index Index.xlsx"),
            SeriesConfig("ten_year_yield", "FULL Nexus Data/US Generic Govt 10 Yr - USGG10YR Index Index.xlsx"),
            SeriesConfig("breakeven_10y", "FULL Nexus Data/BE 10 Year - USGGBE10 Index Index.xlsx"),
            SeriesConfig("fed_funds", "FULL Nexus Data/Fed Funds Target Rate US - FDTR Index Index.xlsx"),
            SeriesConfig("dollar_index", "FULL Nexus Data/DOLLAR INDEX SPOT - DXY Index Index.xlsx"),
            SeriesConfig("economic_surprise", "FULL Nexus Data/Citi Economic Surprise - Unite - CESIUSD INDEX Index.xlsx"),
            SeriesConfig("move_index", "FULL Nexus Data/ICE BofA MOVE Index - MOVE Index Index.xlsx"),
        ],
        "behavioral": [
            SeriesConfig("vix", "FULL Nexus Data/Cboe Volatility Index - VIX Index Index.xlsx"),
            SeriesConfig("move", "FULL Nexus Data/ICE BofA MOVE Index - MOVE Index Index.xlsx"),
            SeriesConfig("sentiment", "FULL Nexus Data/Univ. of Michigan Sentiment - CONSSENT INDEX Index.xlsx"),
            SeriesConfig("confidence", "FULL Nexus Data/Confidence - CONCCONF INDEX Index.xlsx"),
        ],
    }


def run_full_workflow(data_root: Path) -> Dict[str, object]:
    config = default_config(data_root)

    assets = load_series_frame(config["assets"], data_root)
    macro = load_series_frame(config["macro"], data_root)
    behavioral = load_series_frame(config["behavioral"], data_root)

    returns = compute_log_returns(assets)
    macro_changes = compute_feature_changes(macro)

    macro_models, macro_beta_table, macro_corr_table = run_macro_regressions(returns, macro_changes)

    influence_candidates = ["us_equity", "investment_grade", "high_yield", "crypto", "reit"]
    influence_assets = [col for col in influence_candidates if col in returns.columns]
    if len(influence_assets) >= 2:
        cross_asset = estimate_cross_asset_influence(returns[influence_assets])
    else:
        cross_asset = {"lead_lag_table": pd.DataFrame(), "top_leads": pd.DataFrame()}

    regimes = detect_market_regimes(returns["us_equity"])

    behavioral_analysis = behavioral_sentiment_analysis(
        returns["us_equity"], behavioral
    )

    return {
        "asset_prices": assets,
        "asset_returns": returns,
        "macro_levels": macro,
        "macro_changes": macro_changes,
        "macro_models": macro_models,
        "macro_beta_table": macro_beta_table,
        "macro_corr_table": macro_corr_table,
        "cross_asset": cross_asset,
        "regimes": regimes,
        "behavioral": behavioral_analysis,
    }


def format_macro_summary(macro_models: Dict[str, Dict[str, pd.DataFrame]]) -> str:
    if not macro_models:
        return "No macro regressions available."
    lines = []
    for asset, stats in macro_models.items():
        coef = stats["coef"].drop(index="const", errors="ignore").sort_values("coef", ascending=False)
        top = coef.head(5)
        lines.append(f"Asset: {asset}")
        for factor, row in top.iterrows():
            t_val = stats["tvalue"].loc[factor, "tvalue"]
            p_val = stats["pvalue"].loc[factor, "pvalue"]
            lines.append(
                f"  {factor}: coef={row['coef']:.4f}, t={t_val:.2f}, p={p_val:.3f}"
            )
        rsq = stats["rsquared"].iloc[0, 0]
        lines.append(f"  R^2: {rsq:.3f}")
    return "\n".join(lines)


def format_behavioral_summary(behavioral: Dict[str, pd.DataFrame]) -> str:
    summary_df = behavioral.get("summary", pd.DataFrame())
    if summary_df.empty:
        return "No behavioural signal statistics available."
    summary = summary_df.droplevel(1)
    lines = []
    for factor, row in summary.iterrows():
        lines.append(
            f"{factor}: fwd={row['avg_fwd_ret']:.4%}, high={row['avg_high_sentiment_ret']:.4%},"
            f" low={row['avg_low_sentiment_ret']:.4%}, corr={row['corr_with_ret']:.2f}"
        )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Run market dynamics analysis workflow")
    parser.add_argument(
        "--data-root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Path to repository root containing the 'FULL Nexus Data' folder",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="Logging level",
    )
    parser.add_argument(
        "--show-macro",
        action="store_true",
        help="Display the unprocessed macro series levels in the console",
    )
    parser.add_argument(
        "--macro-rows",
        type=int,
        default=10,
        metavar="N",
        help="Number of rows to display when showing macro series (default: 10)",
    )
    parser.add_argument(
        "--macro-series",
        nargs="*",
        default=None,
        help="Optional subset of macro series names to display",
    )
    args = parser.parse_args()

    logging.basicConfig(level=getattr(logging, args.log_level))

    results = run_full_workflow(args.data_root)

    logging.info("Loaded %d asset series", results["asset_prices"].shape[1])
    logging.info("Macro regression summary:\n%s", format_macro_summary(results["macro_models"]))
    logging.info("Behavioral signals:\n%s", format_behavioral_summary(results["behavioral"]))
    top_leads = results["cross_asset"].get("top_leads", pd.DataFrame())
    if isinstance(top_leads, pd.DataFrame) and not top_leads.empty:
        logging.info("Top lead/lag pairs:\n%s", top_leads.to_string(index=False))
    regimes_df = results.get("regimes", pd.DataFrame())
    latest_label = (
        regimes_df["regime_label"].dropna().iloc[-1]
        if isinstance(regimes_df, pd.DataFrame)
        and "regime_label" in regimes_df.columns
        and not regimes_df["regime_label"].dropna().empty
        else "N/A"
    )
    logging.info("Latest regime label: %s", latest_label)

    if args.show_macro:
        macro_levels = results.get("macro_levels", pd.DataFrame())
        if not isinstance(macro_levels, pd.DataFrame) or macro_levels.empty:
            logging.warning("Macro series data unavailable to display.")
        else:
            to_show = macro_levels.copy()
            if args.macro_series:
                requested = [col for col in args.macro_series if col in to_show.columns]
                missing = sorted(set(args.macro_series) - set(requested))
                if missing:
                    logging.warning("Requested macro series missing: %s", ", ".join(missing))
                if requested:
                    to_show = to_show[requested]
                else:
                    logging.warning("No valid macro series selected; showing all available.")
            rows = max(args.macro_rows, 1)
            print("\n=== Macro Series (unprocessed levels) ===")
            print(to_show.head(rows).to_string())


if __name__ == "__main__":
    main()
