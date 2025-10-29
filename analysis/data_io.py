"""Data loading helpers for Nexus analysis notebooks."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import polars as pl
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

# Repository-relative directories used across the analysis workflow.
RAW_DATA_DIR = Path("FULL Nexus Data")
PREPROCESSED_DIR = Path("analysis/preprocessed")
PREPROCESSED_DIR.mkdir(parents=True, exist_ok=True)

# Column aliases to reconcile heterogeneous Bloomberg exports.
SCHEMA_ALIASES: Dict[str, Tuple[Tuple[str, ...], ...]] = {
    "date": (("date", "timestamp", "time", "datetime"),),
    "open": (("open", "px_open", "o"),),
    "high": (("high", "px_high", "h"),),
    "low": (("low", "px_low", "l"),),
    "close": (("close", "px_last", "last", "price"),),
    "adj_close": (("adj close", "adjusted close"),),
    "volume": (("volume", "vol", "qty", "turnover"),),
}

__all__ = [
    "RAW_DATA_DIR",
    "PREPROCESSED_DIR",
    "SCHEMA_ALIASES",
    "_first_present",
    "_coerce_datetime",
    "winsorize_series",
    "load_excel_to_polars",
]


def _first_present(columns: Iterable[str], *candidates_groups: Iterable[str]) -> Optional[str]:
    """Return the first column whose lowered form appears in the candidate groups."""
    cols = list(columns)
    lowered = [c.lower().strip() for c in cols]
    for group in candidates_groups:
        for candidate in group:
            key = candidate.lower().strip()
            if key in lowered:
                return cols[lowered.index(key)]
    return None


def _coerce_datetime(values: List[object]) -> List[Optional[datetime]]:
    """Best-effort conversion of heterogeneous Excel date representations."""
    out: List[Optional[datetime]] = []
    for value in values:
        if value is None:
            out.append(None)
            continue
        if isinstance(value, datetime):
            out.append(value.replace(tzinfo=None))
            continue
        if isinstance(value, date):
            out.append(datetime.combine(value, datetime.min.time()))
            continue
        if isinstance(value, (int, float)):
            try:
                out.append(from_excel(value))
                continue
            except Exception:
                pass
        try:
            parsed_dt = parser.parse(str(value))
            out.append(parsed_dt.replace(tzinfo=None))
        except Exception:
            out.append(None)
    return out


def winsorize_series(series: pl.Series, lower_q: float = 0.01, upper_q: float = 0.99) -> pl.Series:
    """Clip a Polars series to the given quantile bounds."""
    clean = series.drop_nulls()
    if clean.is_empty():
        return series
    lower = clean.quantile(lower_q)
    upper = clean.quantile(upper_q)
    return series.clip(lower, upper)


def _normalise_header_label(raw: str, default: str) -> str:
    """Return a cleaned, snake_case-ish column name with sensible fallbacks."""
    label = (raw or "").strip()
    if not label:
        label = default
    label = re.sub(r"\s+", "_", label)
    label = re.sub(r"[^0-9a-zA-Z_]+", "_", label)
    label = label.strip("_").lower()
    return label or default


def load_excel_to_polars(path: Path) -> pl.DataFrame:
    """Read a flat Bloomberg-exported workbook with openpyxl and return a DataFrame."""
    wb = load_workbook(path, data_only=False, read_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pl.DataFrame()

    n_cols = max(len(row) for row in rows)
    # Attempt to infer field names from the BDH/BDP formula embedded in row 2.
    formula_row = rows[1] if len(rows) > 1 else ()
    formula = formula_row[0] if formula_row else None
    field_names: list[str] = []
    if isinstance(formula, str) and formula.startswith("=_xll"):
        parts = formula.split('"')
        if len(parts) >= 4:
            raw_fields = parts[3]
            field_names = [f.strip() for f in raw_fields.split(",") if f.strip()]

    resolved_header: list[str] = []
    seen: set[str] = set()
    for idx in range(n_cols):
        if idx == 0:
            base_name = "date"
        elif idx - 1 < len(field_names):
            base_name = field_names[idx - 1]
        else:
            header_val = rows[0][idx] if idx < len(rows[0]) else None
            base_name = str(header_val) if header_val is not None else ""
        candidate = _normalise_header_label(base_name, f"col_{idx}")
        suffix = 1
        unique = candidate
        while unique in seen:
            unique = f"{candidate}_{suffix}"
            suffix += 1
        seen.add(unique)
        resolved_header.append(unique)

    data_rows: list[list[object]] = []
    for row in rows[1:]:
        # Drop the Bloomberg formula row and any fully empty rows.
        first_cell = row[0] if row else None
        if isinstance(first_cell, str) and first_cell.startswith("=_xll"):
            continue
        padded = list(row) + [None] * (n_cols - len(row))
        if not any(cell is not None for cell in padded):
            continue
        data_rows.append(padded[:n_cols])

    if not data_rows:
        return pl.DataFrame({name: [] for name in resolved_header})

    columns = {resolved_header[i]: [row[i] for row in data_rows] for i in range(n_cols)}
    return pl.DataFrame(columns)
